import os
import datetime
from openpyxl import load_workbook
from reportlab.pdfgen import canvas
from reportlab.lib.colors import pink, black, red, blue, green
from pdfrw import PdfReader
from pdfrw.buildxobj import pagexobj
from pdfrw.toreportlab import makerl

import s_mail

# Ruta de donde se lee la BBD de excell y la plantilla pdf
files_path = "files" + os.sep
 
# Nombre del archivo excell de donde se leeran los datos
excell_file_name = 'SHAMPOO COLAGENO.xlsx'

# Nombre del archivo pdf que se utiliza como plantilla
pdf_template_file_name = 'template_sh_fases.pdf'

# Ruta donde se guaradaran los archivos
saved_coa_path = ".coas_por_enviar" + os.sep

# Carga del libro excell en wb
wb = load_workbook(filename=files_path + excell_file_name, data_only=True)

# Carga de la hoja de excell que se quiere leer en sheet
sheet = wb['SH. COLAGENO']

# Se inicializa la variable inch la cual se utiliza para los calculo de posicionamiento del texto en el pdf cambas
inch = 72

# Se asigna a col_status la columna que identifica si se ha realizado el COA o no
col_status = sheet['R']

# Se lee la comumna Status y se obtienen los valores de las filas para los que esta columna tiene valor "Nuevo"
for cell in col_status:
    if cell.value == 'N':
        row_number = cell.row
        col_number = cell.col_idx
        sheet.cell(row=row_number, column=col_number, value='V')
        row = sheet[row_number]
        for cell in row:
            if cell.column == 'A':
                fecha_edicion = cell.value.date().strftime("%Y/%m/%d")
            elif cell.column == 'B':
                referencia = cell.value
            elif cell.column == 'C':
                codigo_cliente = cell.value
            elif cell.column == 'D':
                lote = cell.value
            elif cell.column == 'E':
                fecha_elab = cell.value.date().strftime("%Y/%m/%d")
            elif cell.column == 'F':
                fecha_exp = cell.value.date().strftime("%Y/%m/%d")
            elif cell.column == 'G':
                cant_cajas = cell.value
            elif cell.column == 'H':
                und_caja = cell.value
            elif cell.column == 'I':
                total_und = cell.value
            elif cell.column == 'L':
                pH = cell.value
            elif cell.column == 'N':
                viscosidad = cell.value
            elif cell.column == 'P':
                peso_neto = cell.value
            elif cell.column == 'Q':
                coa_number = cell.value
        
        # Probando el loop
        print(referencia + lote)

        # Eliminar esto cuando se cargue estos datos del excell
        cliente = 'Quala'
        aspecto = 'Cumple'
        color = 'Cumple'
        olor = 'Cumple'
        sp_fases = 'Cumple'
        atr_calidad = 'Cumple'


        # Una vez obtenido los valores se imprimen el el pdf con el siguiente código

        created_pdf_name = saved_coa_path + referencia + lote + '.pdf'
        c = canvas.Canvas(created_pdf_name)
        template = PdfReader(files_path + pdf_template_file_name, decompress=False).pages
        t = pagexobj(template[0])
        c.doForm(makerl(c, t))
        c.setFillColor(blue)
        c.drawString(1.8*inch, 8.92*inch, cliente)
        c.drawString(2.42*inch, 8.72*inch, 'COA-No-'+coa_number)
        c.drawString(2.06*inch, 8.51*inch, referencia)
        c.drawString(1.60*inch, 8.32*inch, lote)
        c.drawString(2.84*inch, 8.12*inch, fecha_elab)
        c.drawString(2.75*inch, 7.92*inch, fecha_exp)
        c.drawString(2.50*inch, 7.71*inch, str(cant_cajas))
        c.drawString(2.60*inch, 7.52*inch, und_caja)
        c.drawString(2.58*inch, 7.31*inch, str(total_und))
        c.drawString(6.23*inch, 6.61*inch, aspecto)
        c.drawString(6.23*inch, 6.25*inch, color)
        c.drawString(6.23*inch, 6.06*inch, olor)
        c.drawString(6.23*inch, 5.83*inch, str(pH))
        c.drawString(6.23*inch, 5.63*inch, str(viscosidad))
        c.drawString(6.23*inch, 5.41*inch, str(peso_neto))
        c.drawString(6.23*inch, 5.20*inch, sp_fases)
        c.drawString(6.23*inch, 4.90*inch, atr_calidad)
        c.drawString(1.72*inch, 1.52*inch, fecha_edicion)
        c.drawString(4.85*inch, 1.52*inch, fecha_edicion)


        c.showPage()
        c.save()

wb.save(files_path + excell_file_name)

# Bloque de codigo que prueba la ejecución 
# del envio de los COAs y su gestion
smtp_email = 'yamenmarquez@gmail.com'
smtp_password = 'Lgoogleenon100184'
mail_recipients = ['kpisthatmatter@gmail.com', 'maritzahechavarriaduran@gmail.com']
files_to_attach_path = ".coas_por_enviar" +os.sep
coas_dir_path = "COAs" + os.sep

files_to_attach = []
files = os.listdir(files_to_attach_path)
for f in files:
    if f != '.gitignore':
        files_to_attach.append(f)

# for f in files_to_attach:
#     print(f)

s_mail.send_email(smtp_email, smtp_password, mail_recipients, files_to_attach, files_to_attach_path)

current_path = os.getcwd() + os.sep

for file in files_to_attach:
    origin = current_path + files_to_attach_path + file
    destination = current_path + coas_dir_path + file

    if os.path.exists(origin):
        s_mail.shutil.move(origin, destination)
        print('El archivo ha sido movido a', origin)
    else:
        print('No existe archivo para mover')