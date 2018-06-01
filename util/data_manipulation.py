import os
import datetime
from openpyxl import load_workbook
from constants import COLUMN_NAMES_CONST, column_names_var


# Variables de prueba para testear el script

excell_file_path = "../files"  + os.sep
excell_file_name = "SHAMPOO COLAGENO.xlsx"
excell_sheet_name = 'SH. COLAGENO'

def excell_data(excell_file_path, excell_file_name, excell_sheet_name):
    """Lee y modifica datos de un libro de excell que pose una estructura conocida. Devuelve una
       lista anidada de diccionarios con la estructura de column_names_var from constants

    Argumentos:
    excell_file_path --- Ruta relativa del directotio que contiene el archivo excell qe se quiere manipula
    excell_file_name --- Nombre del archivo excell que se quiere manipular
    excell_sheet_name --- NOmbre de la hoja de excell que contiene los datos que se quieren manipular
    """
    wb = load_workbook(filename=excell_file_path + excell_file_name, data_only=True) # Carga del libro excell en wb

    sheet = wb[excell_sheet_name]

    first_row = sheet['1']

    # count = 0 # Línea para probar la ejecuión correcta de este script

    for cell in first_row:
        # count = count + 1 # Línea para probar la ejecuión correcta de este script
        if  cell.value == 'Status':
            status_column_letter = cell.column
    #     print('Coun->', count) # Línea para probar la ejecuión correcta de este script
    # print(col_status_letter) # Línea para probar la ejecuión correcta de este script

    status_column = sheet[status_column_letter]

    nested_list = []

    # Se lee la comumna Status y se obtienen los valores de las filas para los que esta columna tiene valor "Nuevo"
    for cell in status_column:
        if cell.value == 'N':
            # count = count + 1
            row_number = cell.row
            col_number = cell.col_idx
            sheet.cell(row=row_number, column=col_number, value='N')
            row = sheet[row_number]
            for cell in row:
                for k, v in COLUMN_NAMES_CONST.items():
                    if v == sheet[str(cell.column)+'1'].value: 
                        # print(count)
                        if isinstance(cell.value, datetime.datetime):
                            column_names_var[k] = cell.value.date().strftime("%Y/%m/%d")
                        else:
                            column_names_var[k] = cell.value
                    else:
                        column_names_var[k] = 'empty'

            
                        # print("%s,%s" % (k, v))
                #         COLUMN_NAMES[k] = cell.value
                # print(count)
                # print(COLUMN_NAMES[k])
                
            nested_list.append(column_names_var)
    print(nested_list)

    wb.save(excell_file_path + excell_file_name)

    return nested_list

excell_data(excell_file_path, excell_file_name, excell_sheet_name) # Línea para probar la ejecuión correcta de este script

