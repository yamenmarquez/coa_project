import datetime
import openpyxl
from openpyxl import load_workbook
from reportlab.pdfgen import canvas

from pdfrw import PdfReader
from pdfrw.buildxobj import pagexobj
from pdfrw.toreportlab import makerl
 
# producto = 'producto'
# lote = 'lote'
# cantidad_producida = 0
# viscosidad = 0
# pH = 0
# color = 'color'
# olor = 'olor'
# fecha = 'fecha'

wb = load_workbook(filename = 'coa.xlsx')

sheet_coadb = wb['coa_db']

inch = 72

# for row in sheet_coadb.iter_rows(min_row=1, max_col=8, max_row=3):
#     for cell in row:
#         print(cell.col_idx)
#         print(cell.row)
#         print(cell.value)

col_status = sheet_coadb['I']

for cell in col_status:
    if cell.value == 'Nuevo':
        # print(cell.value)
        # print(cell.row
        row_number = cell.row
        col_number = cell.col_idx
        # sheet_coadb.cell(row=row_number, column=col_number, value='Viejo')
        row = sheet_coadb[row_number]
        for cell in row:
            if cell.column == 'A':
                producto = cell.value
            elif cell.column == 'B':
                lote = cell.value
            elif cell.column == 'C':
                cantidad_producida = cell.value
            elif cell.column == 'D':
                viscosidad = cell.value
            elif cell.column == 'E':
                pH = cell.value
            elif cell.column == 'F':
                color = cell.value
            elif cell.column == 'G':
                olor = cell.value
            elif cell.column == 'H':
                fecha = cell.value.date().strftime("%Y/%m/%d")
    
        # print(producto)
        # print(lote)
        # print(cantidad_producida)
        # print(viscosidad)
        # print(pH)
        # print(color)
        # print(olor)
        # print(fecha)
        pdf_name = producto + lote + '.pdf'
        c = canvas.Canvas(pdf_name)
        template = PdfReader('template_sh_fases.pdf',decompress=False).pages
        t = pagexobj(template[0])
        c.doForm(makerl(c, t))
        c.drawString(1.8*inch, 8.91*inch, producto)
        c.showPage()
        c.save()


                
            # if cell.is_date:
            #     print(cell.value.date().strftime("%Y/%m/%d"))
            # else:
            #     print(cell.value)
wb.save('coa.xlsx')
